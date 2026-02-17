"""
Strukturis Pro — Parser de Dados e Exportação Inteligente
Exporta para Excel (formatado), CSV (UTF-8 BOM), e PDF (relatório).
"""

import pandas as pd
import re
from io import StringIO, BytesIO
from datetime import datetime


class DataParser:
    @staticmethod
    def parse_to_dataframe(text):
        """
        Converte texto não-estruturado em DataFrame.
        Estratégia: Procura linhas com estrutura similar (delimitadores).
        """
        lines = text.split('\n')
        data = []

        for line in lines:
            line = line.strip()
            if not line:
                continue
            # Divide por 2+ espaços ou tabs
            parts = re.split(r'\s{2,}|\t', line)
            if len(parts) > 1:
                data.append(parts)

        if not data:
            return pd.DataFrame([text], columns=["Conteúdo"])

        # Normaliza número de colunas
        max_cols = max(len(row) for row in data)
        for row in data:
            while len(row) < max_cols:
                row.append('')

        df = pd.DataFrame(data)

        # Tenta promover 1ª linha como cabeçalho
        first_row = df.iloc[0]
        is_header = all(
            not re.match(r'^[\d.,]+$', str(v).strip())
            for v in first_row if str(v).strip()
        )
        if is_header and len(df) > 1:
            df.columns = [str(v).strip() for v in first_row]
            df = df.iloc[1:].reset_index(drop=True)
        else:
            df.columns = [f"Coluna {i+1}" for i in range(df.shape[1])]

        return df


class Exporter:
    """Exportação inteligente com formatação profissional."""

    @staticmethod
    def to_excel(df, path, doc_title="Strukturis Pro", metadata=None):
        """
        Exporta DataFrame para Excel com formatação profissional:
        - Cabeçalho estilizado com cores
        - Largura automática de colunas
        - Filtros automáticos
        - Rodapé com data/hora
        """
        try:
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                # Escreve dados a partir da linha 3 (reserva linhas para título)
                start_row = 2
                df.to_excel(writer, index=False, sheet_name='Dados', startrow=start_row)

                ws = writer.sheets['Dados']

                # Título
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(df.shape[1], 3))
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = doc_title
                title_cell.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
                title_cell.fill = PatternFill(start_color='007ACC', end_color='007ACC', fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Metadata row
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(df.shape[1], 3))
                meta_cell = ws.cell(row=2, column=1)
                meta_text = f"Gerado por Strukturis Pro em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                if metadata:
                    meta_text += f" | {metadata}"
                meta_cell.value = meta_text
                meta_cell.font = Font(name='Segoe UI', size=9, italic=True, color='666666')
                meta_cell.alignment = Alignment(horizontal='center')

                # Header styling
                header_fill = PatternFill(start_color='2D2D30', end_color='2D2D30', fill_type='solid')
                header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
                thin_border = Border(
                    left=Side(style='thin', color='3E3E42'),
                    right=Side(style='thin', color='3E3E42'),
                    top=Side(style='thin', color='3E3E42'),
                    bottom=Side(style='thin', color='3E3E42'),
                )

                for col_idx in range(1, df.shape[1] + 1):
                    cell = ws.cell(row=start_row + 1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                # Data styling
                data_font = Font(name='Segoe UI', size=10)
                alt_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

                for row_idx in range(start_row + 2, start_row + 2 + df.shape[0]):
                    for col_idx in range(1, df.shape[1] + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = data_font
                        cell.border = thin_border
                        if (row_idx - start_row) % 2 == 0:
                            cell.fill = alt_fill

                # Auto-width
                for col_idx in range(1, df.shape[1] + 1):
                    max_len = max(
                        len(str(ws.cell(row=r, column=col_idx).value or ''))
                        for r in range(start_row + 1, start_row + 2 + df.shape[0])
                    )
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

                # Auto-filter
                ws.auto_filter.ref = ws.dimensions

        except Exception as e:
            # Fallback simples
            df.to_excel(path, index=False)

    @staticmethod
    def to_csv(df, path, metadata=None):
        """
        Exporta para CSV com encoding UTF-8 BOM e separador ';' (padrão BR).
        """
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            # Metadados no cabeçalho
            f.write(f"# Strukturis Pro - Exportação {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
            if metadata:
                f.write(f"# {metadata}\n")
            f.write("#\n")
            df.to_csv(f, index=False, sep=';')

    @staticmethod
    def to_pdf_report(df, path, title="Relatório Strukturis Pro", metadata=None):
        """
        Gera PDF relatório formatado com tabela estilizada.
        Usa reportlab se disponível, senão fallback com PyMuPDF.
        """
        try:
            return Exporter._pdf_via_reportlab(df, path, title, metadata)
        except ImportError:
            return Exporter._pdf_via_fitz(df, path, title, metadata)

    @staticmethod
    def _pdf_via_reportlab(df, path, title, metadata):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm

        # Decide orientation based on columns
        page_size = landscape(A4) if df.shape[1] > 5 else A4

        doc = SimpleDocTemplate(path, pagesize=page_size,
                                leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                                topMargin=1.5 * cm, bottomMargin=1.5 * cm)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#007ACC'),
            spaceAfter=10,
        )
        elements.append(Paragraph(title, title_style))

        # Metadata
        meta_style = ParagraphStyle(
            'Meta',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=20,
        )
        meta_text = f"Gerado por Strukturis Pro em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        if metadata:
            meta_text += f"<br/>{metadata}"
        elements.append(Paragraph(meta_text, meta_style))

        # Table
        table_data = [list(df.columns)] + df.values.tolist()

        # Truncar textos longos
        for i, row in enumerate(table_data):
            table_data[i] = [str(v)[:60] if len(str(v)) > 60 else str(v) for v in row]

        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D2D30')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#3E3E42')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)

        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        elements.append(Paragraph("Strukturis Pro © — Análise Inteligente de Documentos", footer_style))

        doc.build(elements)
        return True

    @staticmethod
    def _pdf_via_fitz(df, path, title, metadata):
        """Fallback usando PyMuPDF para gerar PDF simples."""
        import fitz

        doc = fitz.open()
        page = doc.new_page(width=842, height=595)  # A4 Landscape

        # Title
        page.insert_text((30, 40), title, fontsize=18, color=(0, 0.478, 0.8))

        # Meta
        meta = f"Gerado por Strukturis Pro em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        page.insert_text((30, 60), meta, fontsize=9, color=(0.5, 0.5, 0.5))

        # Table as text
        y = 90
        # Header
        header = " | ".join(str(c) for c in df.columns)
        page.insert_text((30, y), header, fontsize=9, color=(0, 0, 0))
        y += 5
        page.draw_line((30, y), (812, y), color=(0.2, 0.2, 0.2), width=1)
        y += 15

        for _, row in df.iterrows():
            if y > 560:
                page = doc.new_page(width=842, height=595)
                y = 30
            line = " | ".join(str(v)[:40] for v in row.values)
            page.insert_text((30, y), line, fontsize=8, color=(0.1, 0.1, 0.1))
            y += 14

        doc.save(path)
        doc.close()
        return True

    @staticmethod
    def to_txt(text, path):
        """Exporta texto puro."""
        with open(path, 'w', encoding='utf-8') as f:
            f.write(text)
